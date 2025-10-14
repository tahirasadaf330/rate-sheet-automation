import pandas as pd, os, re, numpy as np
from datetime import datetime
from dateutil import parser as dparse
from openpyxl import load_workbook
import re, unicodedata

REQUIRED_COLS = [
    'Dst Code', 'Rate', 'Effective Date', 'Billing Increment'
]

ROW_FALLBACK_THRESHOLD = 100  # trigger pandas fallback if streaming returns < 100 rows


BILLING_PAIRS = [
    ('initial_period', 'recurring_period'),
    ('initial_period', 'subsequent_period'),
    ('initial_increment', 'next_increment'),
    ('min_bill', 'billing_step'),
    ('first_increment', 'second_increment'),
    ('first_increment', 'additional_increment'),
]

EXCEL_EPOCH = datetime(1899, 12, 30)          # Excel’s epoch (PC versions)

DEBUG = True
def dbg(*a, **k):
    if DEBUG:
        print(*a, **k)

def _codepoints(s):
    s = str(s)
    return " ".join(f"U+{ord(ch):04X}" for ch in s)


# ─────────────────────────── Dst Code expander ───────────────────────────

_DASHES_RE = re.compile(r'[–—−]')  # normalize en/em/minus to simple "-"

def _parse_dst_code_list(code_str: str) -> list[str]:
    """Split 'Dst Code' strings on comma/semicolon; expand hyphen ranges inclusive.
       Treat everything that isn't a digit or '-' as noise. Preserve zero padding."""
    if pd.isna(code_str):
        return []
    s = str(code_str).strip()
    if not s:
        return []

    s = _DASHES_RE.sub('-', s)  # unify weird dashes
    parts = re.split(r'[;,]', s)  # comma/semicolon = separate codes

    out: list[str] = []
    for part in parts:
        token = re.sub(r'[^0-9\-]', '', part.strip())  # keep digits and hyphen only
        if not token:
            continue

        if '-' in token:
            lo_str, hi_str = token.split('-', 1)
            if lo_str.isdigit() and hi_str.isdigit():
                lo, hi = int(lo_str), int(hi_str)
                if hi < lo:  # be forgiving if they reversed it
                    lo, hi = hi, lo
                width = max(len(lo_str), len(hi_str))
                out.extend(f"{n:0{width}d}" for n in range(lo, hi + 1))
                continue

        # single code
        out.append(token)

    return out

def expand_dst_code_rows(df: pd.DataFrame) -> pd.DataFrame:
    """Explode rows that have multiple/ranged DST codes into one row per code."""
    out = df.copy()
    out['Dst Code'] = out['Dst Code'].astype(str).str.strip()
    out['__dst_list'] = out['Dst Code'].apply(_parse_dst_code_list)
    out = out.explode('__dst_list', ignore_index=True)
    out['Dst Code'] = out['__dst_list'].fillna('')
    out = out[out['Dst Code'] != ''].drop(columns='__dst_list').reset_index(drop=True)
    return out

# ---- Header pre-cleaner: strip currency symbols/codes and other noise ----

_CURRENCY_SYMBOLS_RE = r'[$£€¥₹₩₽₺₫₪₴₦₱₲₵₡₭฿₮₸₼]'
_CURRENCY_CODES = (
    "usd|eur|gbp|jpy|cny|rmb|cad|aud|nzd|inr|chf|sar|aed|rub|brl|mxn|zar|"
    "sek|nok|dkk|pln|huf|try|ils|krw|idr|myr|thb|php|vnd|uah|ron|czk|ars|"
    "clp|cop|pen|twd|hkd|sgd|ngn|kzt"
)
_CURRENCY_CODES_RE = re.compile(rf"(?i)\b(?:{_CURRENCY_CODES})\b")
_CURRENCY_SYMBOLS_RE = re.compile(r"[$£€¥₹₩₽₺₫₪₴₦₱₲₵₡₭฿₮₸₼]")

def _preclean_header_token(s: str) -> str:
    """Strip currency symbols/codes and common junk from HEADER labels."""
    s = str(s).replace("\n", " ")
    s = _CURRENCY_SYMBOLS_RE.sub(" ", s)          # $, €, etc.
    s = re.sub(r"\(.*?\)", " ", s)                # drop parentheticals like (USD)
    s = _CURRENCY_CODES_RE.sub(" ", s)            # USD, EUR, etc.
    s = re.sub(r"(?i)\bper\s*(min(?:ute)?|sec(?:ond)?)\b", " ", s)
    s = re.sub(r"[/\\|:]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def _strip_currency_words_from_key(key: str) -> str:
    """After _norm(), remove currency tokens that became words (e.g. rate_usd)."""
    # kill leading/trailing or middle _usd/_eur tokens
    key = re.sub(rf"(?:^|_)(?:{_CURRENCY_CODES})(?=_|$)", "", key, flags=re.I)
    key = re.sub(r"_+", "_", key).strip("_")
    return key

#________________________________handeling seperate/duplicate billing increment columns ──────────────────────────────────

def _coalesce_billing_increment_dupes(df: pd.DataFrame, col_name: str = 'Billing Increment') -> pd.DataFrame:
    """
    If there are 2+ columns named `col_name`, combine the FIRST TWO into a single 'a/b' string per row:
      - a = last integer from the first dup column
      - b = last integer from the second dup column
      - If only one side exists on a row, duplicate it as a/a or b/b
    Drops all original dup columns and inserts a single consolidated column at the
    position of the first duplicate. No-op if <2 duplicates exist.
    """
    # locate duplicates by position, not by label
    mask = (df.columns == col_name)
    idxs = np.flatnonzero(mask)
    if len(idxs) < 2:
        return df  # nothing to do

    # grab the first two duplicate columns by position
    block = df.iloc[:, idxs]
    left_raw = block.iloc[:, 0]
    right_raw = block.iloc[:, 1]

    # reuse your existing number extractor
    def _num_last(x) -> str:
        return _last_num(x)  # already defined in your file

    left = left_raw.map(_num_last).fillna('')
    right = right_raw.map(_num_last).fillna('')

    # build the consolidated a/b values
    out = pd.Series('', index=df.index, dtype='object')
    both = (left != '') & (right != '')
    only_l = (left != '') & (right == '')
    only_r = (left == '') & (right != '')

    out.loc[both] = left[both] + '/' + right[both]
    out.loc[only_l] = left[only_l] + '/' + left[only_l]
    out.loc[only_r] = right[only_r] + '/' + right[only_r]
    # rows where both sides empty remain ''

    # drop all dup columns, insert the consolidated one at the original first position
    first_pos = int(idxs[0])
    df = df.drop(columns=block.columns)
    df.insert(min(first_pos, len(df.columns)), col_name, out)

    return df


def _last_num(s: str) -> str:
    m = re.findall(r'\d+', str(s))
    return m[-1] if m else ''

def _synthesize_billing_increment(df: pd.DataFrame) -> pd.DataFrame:

    df = _coalesce_billing_increment_dupes(df)

    # If already present and non-empty anywhere, keep it
    if 'Billing Increment' in df.columns and df['Billing Increment'].astype(str).str.strip().ne('').any():
        return df
    
    print("\n\n\nSynthesizing 'Billing Increment' from other columns...\n\n\n")

    # map normalized -> actual column name
    norm2real = { _norm(c): c for c in df.columns }

    pairs = [
        ('initial_period', 'recurring_period'),
        ('initial_period', 'subsequent_period'),
        ('initial_increment', 'next_increment'),
        ('min_bill', 'billing_step'),
        ('first_increment', 'second_increment'),
        ('first_increment', 'additional_increment'),
    ]

    for a, b in pairs:
        if a in norm2real and b in norm2real:
            ax = df[norm2real[a]].map(_last_num)
            bx = df[norm2real[b]].map(_last_num)
            mask = (ax != '') & (bx != '')
            if mask.any():
                df['Billing Increment'] = np.where(mask, ax + '/' + bx, '')
                return df

    # fallback: single-column duplication if we only have one increment-ish column
    singles = ['initial_period', 'min_bill', 'first_increment']
    for k in singles:
        if k in norm2real:
            one = df[norm2real[k]].map(_last_num)
            if one.ne('').any():
                df['Billing Increment'] = np.where(one != '', one + '/' + one, '')
                return df

    # ensure the column exists so downstream selection doesn't explode
    df['Billing Increment'] = ''
    return df

#________________________________ read raw ──────────────────────────────────

def _raw_from_ws(ws) -> pd.DataFrame:
    try:
        # Try to calculate dimensions without forcing
        ws.calculate_dimension()
    except ValueError:
        # If it fails, use calculate_dimension with force=True
        print("DEBUG: Extraction failed without force; retrying with force=True")
        ws.calculate_dimension(force=True)

    rows = list(ws.iter_rows(values_only=True))
    raw = pd.DataFrame(rows)
    raw.dropna(how="all", inplace=True)
    raw.dropna(axis=1, how="all", inplace=True)
    raw.reset_index(drop=True, inplace=True)
    return raw.astype("string")


# ── PATCH 1: make the pandas reader actually try multiple engines & the target sheet ──
def _raw_from_excel_pandas(path: str, sheet) -> pd.DataFrame:
    """
    Read a sheet using pandas.read_excel with multiple engines.
    Returns a raw grid (no headers assigned), dropping all-empty rows/columns.
    """
    engines = ("calamine", "openpyxl")  # try calamine first; fall back to openpyxl
    last_exc = None
    for eng in engines:
        try:
            df = pd.read_excel(
                path,
                sheet_name=sheet if sheet is not None else 0,
                engine=eng,
                header=None,     # we want a raw grid
                dtype=str
            )
            # If pandas returns a dict (when sheet_name=None), normalize to the first sheet
            if isinstance(df, dict):
                # pick the first item deterministically
                first_key = sorted(df.keys())[0]
                df = df[first_key]

            df = pd.DataFrame(df)
            df.dropna(how="all", inplace=True)
            df.dropna(axis=1, how="all", inplace=True)
            df.reset_index(drop=True, inplace=True)
            return df.astype("string")
        except Exception as e:
            dbg(f"[pandas engine={eng}] read failed: {e}")
            last_exc = e
            continue
    raise last_exc if last_exc else RuntimeError("read_excel failed with both calamine and openpyxl")

# ── PATCH 2: strengthen _read_raw_matrix to try pandas when there are 0 worksheets
#             and as a final fallback when header detection fails on all sheets. ──
def _read_raw_matrix(path: str, sheet=0) -> pd.DataFrame:
    ext = os.path.splitext(path)[1].lower()
    if ext in ('.xlsx', '.xlsm', '.xls'):
        # Optional: route legacy .xls straight to pandas, since openpyxl can't read BIFF .xls
        if ext == '.xls':
            try:
                print("DEBUG: Trying the read excel pandas method (.xls fast path)")
                raw_pd = _raw_from_excel_pandas(path, sheet if sheet is not None else 0)
                print("DEBUG: Gotcha using the read excel pandas method (.xls fast path)")
                return raw_pd
            except Exception:
                # fall through to the openpyxl loop anyway, in case the file is misnamed
                pass

        print("DEBUG: Trying the workbook method")
        wb = load_workbook(path, data_only=True, read_only=True)

        # If there are ZERO normal worksheets (chartsheet-only, corrupted, or placeholder),
        # try a pandas fallback immediately.
        if len(wb.worksheets) == 0:
            dbg("[excel] 0 worksheets detected by openpyxl → trying pandas fallback")
            try:
                raw_pd = _raw_from_excel_pandas(path, sheet if sheet is not None else 0)
                wb.close()
                return raw_pd
            except Exception as e:
                wb.close()
                raise ValueError(f"No worksheets found and pandas fallback failed: {e}")

        # try requested sheet first, then all others
        try_order = []
        if isinstance(sheet, int) and 0 <= sheet < len(wb.worksheets):
            try_order.append(sheet)
        elif isinstance(sheet, str):
            try_order += [i for i, ws in enumerate(wb.worksheets) if ws.title == sheet]
        try_order += [i for i in range(len(wb.worksheets)) if i not in try_order]

        # helper: pandas-all-sheets fallback (even when row count >= threshold)
        def _try_pandas_all_sheets():
            for j in range(len(wb.worksheets)):
                try:
                    raw_pd = _raw_from_excel_pandas(path, j)
                    try:
                        _ = detect_header_row(raw_pd)
                        dbg(f"[excel-fallback] pandas finally found headers on sheet #{j}")
                        return raw_pd
                    except ValueError:
                        continue
                except Exception as e:
                    dbg(f"[excel-fallback] pandas read failed on sheet #{j}: {e}")
            return None

        for i in try_order:
            raw_stream = _raw_from_ws(wb.worksheets[i])

            # Fallback trigger BEFORE header detection if the streaming read seems too short
            chosen = raw_stream
            if raw_stream.shape[0] < ROW_FALLBACK_THRESHOLD:
                try:
                    raw_pd = _raw_from_excel_pandas(path, i)  # same sheet index
                    if raw_pd.shape[0] > raw_stream.shape[0]:
                        dbg(f"[excel-fallback] streaming rows={raw_stream.shape[0]} < {ROW_FALLBACK_THRESHOLD}; "
                            f"pandas rows={raw_pd.shape[0]} -> using pandas for sheet #{i}")
                        chosen = raw_pd
                    else:
                        dbg(f"[excel-fallback] streaming rows={raw_stream.shape[0]} < {ROW_FALLBACK_THRESHOLD}; "
                            f"pandas rows={raw_pd.shape[0]} not larger -> keeping streaming for sheet #{i}")
                except Exception as e:
                    dbg(f"[excel-fallback] pandas read failed on sheet #{i}: {e}")

            try:
                print('\n\nDEBUG: Calling the detect header row function from read_raw_matrix\n\n')
                _ = detect_header_row(chosen)  # will raise if not found
                wb.close()
                print("\n\nDEBUG: Gotcha using the read excel openpyxl method")
                return chosen  # this sheet has the headers; use it
            except ValueError:
                print("\n\nDEBUG: Failed the detect_header_row check")
                # don’t return yet; try other sheets
                continue

        # If we exhausted all openpyxl sheets without finding headers, try pandas across all sheets
        dbg("[excel] openpyxl could not find headers on any sheet → trying pandas across all sheets")
        raw_pd_any = _try_pandas_all_sheets()
        wb.close()
        if raw_pd_any is not None:
            return raw_pd_any

        raise ValueError("No sheet contains all required headers (openpyxl and pandas fallbacks failed).")
    else:
        # Non-Excel → treat as CSV/TSV/etc.
        return pd.read_csv(path, header=None, dtype=str)

# In the detect_header_row function:
def detect_header_row(raw: pd.DataFrame) -> int:
    """
    Find the row index that, after normalization + alias mapping,
    contains ALL required canonical columns, but stops after 1000 rows.
    """
    targets = set(REQUIRED_COLS)

    best_count = -1
    best_row = -1
    best_covered = set()

    # Add a max row limit of 1000
    max_row_limit = 1000

    for idx, row in raw.iterrows():
        if idx >= max_row_limit:
            break  # Stop searching after 1000 rows

        # raw cell texts on this row (skip NaN)
        cells = [x for x in row if pd.notna(x)]
        precleaned = [_preclean_header_token(x) for x in cells]
        normed = [_norm(x) for x in precleaned]
        mapped = [ALIAS_MAP.get(n) or _match_alias_substring(n) for n in normed]

        covered = {m for m in mapped if m}
        
        if idx <=10:
            print(f"DEBUG TARGETS: {targets}")
            print("DEBUG IN THE DETECT HEADER ROW: covered so far", covered)

        ###################
        # ---------- tolerate missing Billing Increment if related headers exist ----------
        if 'Billing Increment' not in covered:
            norm_set = set(normed)

            def _has_key(key: str) -> bool:
                # fuzzy: exact, prefix, suffix, or underscore-delimited infix
                return any(
                    n == key
                    or n.startswith(key + '_')
                    or n.endswith('_' + key)
                    or ('_' + key + '_') in ('_' + n + '_')
                    for n in norm_set
                )

            # Pairs allow perfect synthesis later
            pair_hit = any(_has_key(a) and _has_key(b) for (a, b) in BILLING_PAIRS)

            # Singles are also enough, because _synthesize_billing_increment can duplicate a single
            singles = ['initial_period', 'min_bill', 'first_increment', 'increment']
            single_hit = any(_has_key(k) for k in singles)

            if pair_hit or single_hit:
                covered.add('Billing Increment')

            dbg("  billing_pair_hit:", pair_hit, "singles_hit:", single_hit,
                "pairs_checked:", BILLING_PAIRS, "singles_checked:", singles)
        # ----------------------------------------------------------------------------- 
        ##################

        missing = targets - covered

        # DEBUG dump
        dbg(f"[hdr-scan] row={idx}")
        dbg("  cells:", [repr(x) for x in cells])
        dbg("  precleaned:", [repr(x) for x in precleaned])
        dbg("  normed:", normed)
        dbg("  mapped:", mapped)
        if missing:
            dbg("  still-missing:", sorted(missing))
        else:
            dbg(f"[hdr-scan] FOUND header row -> {idx}")
            return idx

        # track best partial coverage to help when failing
        if len(covered & targets) > best_count:
            best_count = len(covered & targets)
            best_row = idx
            best_covered = covered & targets

    # If we got here, we failed. Print the best candidate with codepoints.
    dbg(f"[hdr-scan] best coverage: {best_count} on row {best_row} -> {sorted(best_covered)}")
    if best_row != -1:
        best_cells = [x for x in raw.iloc[best_row] if pd.notna(x)]
        dbg("[hdr-scan] best row cells (repr):", [repr(x) for x in best_cells])
        dbg("[hdr-scan] best row cells (_norm):", [_norm(x) for x in best_cells])
        dbg("[hdr-scan] best row cells (codepoints):", [_codepoints(str(x)) for x in best_cells])

    raise ValueError(
        "Header not found. None of the rows contained all required columns "
        f"after normalization/aliasing. Required: {REQUIRED_COLS}"
    )


# kill zero-widths and collapse all whitespace/slashes/hyphens; strip punctuation
_ZW = r'[\u200B-\u200D\uFEFF]'

def _norm(s: str) -> str:
    s = unicodedata.normalize('NFKC', str(s)).lower()
    s = re.sub(_ZW, '', s)                 # remove zero-width chars
    s = re.sub(r'[\s/\\\-]+', ' ', s)      # any whitespace, slash, hyphen -> single space
    s = re.sub(r'[^\w ]+', '', s)          # drop punctuation like ., :, (), etc.
    s = ' '.join(s.split())                # collapse to single spaces
    return s.replace(' ', '_')

# Map of normalized header variants -> canonical
ALIAS_MAP = {
    # Dst Code
    'dst_code': 'Dst Code',
    'dstcode': 'Dst Code',
    'code': 'Dst Code',
    'destination_code': 'Dst Code',
    'dest_code': 'Dst Code',
    'area_code': 'Dst Code',
    'dial_code': 'Dst Code',
    'dialcode': 'Dst Code',
    'codes': 'Dst Code',
    'dial_codes': 'Dst Code',
    'area_code': 'Dst Code',


    # Rate
    'rate': 'Rate',
    'rates': 'Rate',
    'price': 'Rate',
    'new_rates': 'Rate',
    'new_rate': 'Rate',
    'cost': 'Rate',
    'pricing': 'Rate',
    'price': 'Rate',
    'pricing_in': 'Rate',
    'rate_min($)': 'Rate',
    'new_price': 'Rate',
    'price_peak': 'Rate',
    'pricemin': 'Rate',

    # Effective Date
    'effective_date': 'Effective Date',
    'effective': 'Effective Date',
    'eff_date': 'Effective Date',
    'effective_from': 'Effective Date',
    'start_date': 'Effective Date',
    'valid_from': 'Effective Date',
    'date': 'Effective Date',
    'effectivedate': 'Effective Date',
    'efective_date': 'Effective Date',  

    # Billing Increment
    'billing_increment': 'Billing Increment',
    'billing_increament': 'Billing Increment',   # common typo
    'billing_increments': 'Billing Increment',
    'billing_inc': 'Billing Increment',
    'billing': 'Billing Increment',
    'billingincrement': 'Billing Increment',
    'rounding_rules': 'Billing Increment',
    'rounding': 'Billing Increment',
    'billing_terms': 'Billing Increment',
    'increment': 'Billing Increment',
}


##############################################
# new code for handeling the junk that compes with the col names

# --- Substring alias matching helpers ---
def _normalize_header_key(s: str) -> str:
    """
    Lowercase, replace non-alphanumerics with underscores, collapse repeats,
    and strip edges. Keeps behavior tight and predictable.
    """
    s = unicodedata.normalize('NFKC', str(s)).lower()
    s = re.sub(r'[^0-9a-z]+', '_', s)
    s = re.sub(r'_+', '_', s).strip('_')
    return s

# Build a normalized alias map once. Keys should already be "alias-style"
# but this makes it robust to future edits.
ALIAS_MAP_NORM = { _normalize_header_key(k): v for k, v in ALIAS_MAP.items() }

def _match_alias_substring(normalized_key: str, alias_map: dict = ALIAS_MAP_NORM):
    """
    Try to map a normalized header by substring match against alias keys.
    Returns canonical header string or None.
    Preference order:
      1) exact match
      2) longest alias that is a substring of the key
    """
    # exact hit first
    if normalized_key in alias_map:
        return alias_map[normalized_key]

    # substring hits, longest alias wins to avoid 'date' beating 'effective_date'
    for alias in sorted(alias_map.keys(), key=len, reverse=True):
        if alias and alias in normalized_key:
            return alias_map[alias]
    return None


#################################################

def _canonicalize_headers(df: pd.DataFrame) -> pd.DataFrame:
    original = list(df.columns)

    # 1) Preclean labels (kill $, €, USD, etc.)
    preclean_map = {c: _preclean_header_token(c) for c in original}
    # 2) Normalize
    norm_map = {c: _norm(preclean_map[c]) for c in original}
    # 3) Strip currency words that survived normalization (e.g., rate_usd -> rate)
    key_map = {c: _strip_currency_words_from_key(norm_map[c]) for c in original}
    # 4) Alias lookup on the final key

    alias_hit = {}
    for c in original:
        # Skip renaming for 'Dst Code Name'
        if c.lower() == 'dst code name':
            alias_hit[c] = 'Dst Code Name'
            continue  # skip this column

        if c.lower() == 'current rate usd':
            alias_hit[c] = 'CURRENT RATE USD'
            continue  # skip this column

        key = key_map[c]  # already precleaned+normalized version of c
        hit = ALIAS_MAP.get(key)
        if not hit:
            # fallback: alias substring match on the normalized key
            hit = _match_alias_substring(key)
        alias_hit[c] = hit

    # DEBUG
    dbg("[canon] original -> preclean -> norm -> key_strip -> alias:")
    for c in original:
        dbg(f"  {repr(c)}  ->  {repr(preclean_map[c])}  ->  {norm_map[c]}  ->  {key_map[c]}  ->  {alias_hit[c]}")
        if not alias_hit[c]:
            dbg("    codepoints(original):", _codepoints(c))

    # If alias matches, use canonical; otherwise keep the cleaned label
    mapped = {c: (alias_hit[c] if alias_hit[c] else preclean_map[c]) for c in original}
    df = df.rename(columns=mapped)

    # ---------- NEW: tolerate missing Billing Increment if a known pair exists ----------
    missing = [c for c in REQUIRED_COLS if c not in df.columns]

    if 'Billing Increment' in missing:
        # normalize the CURRENT df.columns (post-rename) for fuzzy matching
        normed_current = {_norm(_preclean_header_token(c)) for c in df.columns}

        def _has_key(key: str) -> bool:
            # fuzzy: exact, prefix, suffix, or underscore-delimited infix
            return any(
                n == key or
                n.startswith(key + '_') or
                n.endswith('_' + key) or
                ('_' + key + '_') in ('_' + n + '_')
                for n in normed_current
            )

        pair_hit = any(_has_key(a) and _has_key(b) for (a, b) in BILLING_PAIRS)
        dbg("[canon] billing_pair_hit:", pair_hit, "pairs_checked:", BILLING_PAIRS)

        if pair_hit:
            # don’t count BI as missing; create placeholder so later selection won’t crash
            missing = [m for m in missing if m != 'Billing Increment']
            if 'Billing Increment' not in df.columns:
                df['Billing Increment'] = ''   # _synthesize_billing_increment will fill this later
            dbg("[canon] Billing Increment satisfied via header pair; will synthesize values later.")
    # -----------------------------------------------------------------------------------

    # Final guard
    if missing:
        dbg("[canon] df.columns:", list(df.columns))
        dbg("[canon] missing required:", missing)
        dbg("[canon] precleaned originals:", preclean_map)
        dbg("[canon] normalized originals:", norm_map)
        dbg("[canon] stripped keys:", key_map)
        raise ValueError(
            "Missing required columns: "
            f"{missing}. Found headers: {original}. "
            f"Precleaned: {preclean_map}. "
            f"Normalized: {norm_map}. "
            f"Stripped keys: {key_map}. "
            "Add more variants to ALIAS_MAP or harden _norm/_preclean_header_token."
        )
    return df


# ──────────────────────────────── footer ─────────────────────────────────

# Customize the keywords if you want to add more
_NOTES_RE = re.compile(r'^(note|notes|remark|remarks|disclaimer|footer|terms and conditions)\b', re.IGNORECASE)

def _is_empty_row(row: pd.Series) -> bool:
    """True if every cell in the row is NaN or only whitespace."""
    for v in row:
        if pd.isna(v):
            continue
        if isinstance(v, str):
            if v.strip() != "":
                return False
        else:
            # non-string, non-NaN value counts as data
            return False
    return True

def trim_after_notes_and_strip_blank_above(df: pd.DataFrame) -> pd.DataFrame:
    """
    - Find the first row where ANY cell starts with one of the notes keywords.
    - Cut the DataFrame to everything ABOVE that row.
    - Then remove trailing blank rows immediately above the notes marker.
    If no notes marker is found, returns df unchanged.
    """
    if df.empty:
        return df

    # Build a normalized view for matching
    # norm = df.applymap(lambda x: "" if pd.isna(x) else str(x).strip())

    try:
        norm = df.map(lambda x: "" if pd.isna(x) else str(x).strip())
    except AttributeError:
        norm = df.applymap(lambda x: "" if pd.isna(x) else str(x).strip())


    # Row-wise: does ANY cell start with a notes-like keyword?
    marker_mask = norm.apply(lambda r: any(_NOTES_RE.match(cell) for cell in r), axis=1)

    if not marker_mask.any():
        # no notes marker -> return as-is
        return df

    # Position of first notes marker (iloc)
    marker_index_label = marker_mask.idxmax()  # first True by index order
    marker_iloc = df.index.get_loc(marker_index_label)

    # Everything strictly ABOVE the marker
    end_iloc = marker_iloc - 1

    # Strip trailing blank rows immediately above the marker
    while end_iloc >= 0 and _is_empty_row(df.iloc[end_iloc]):
        end_iloc -= 1

    # If nothing left, return empty frame with same columns
    if end_iloc < 0:
        return df.iloc[0:0].copy()

    return df.iloc[:end_iloc + 1].copy()

# ───────────────────────────────── helpers ──────────────────────────────────

def _normalize_excel_writer_path(path: str) -> tuple[str, dict]:
    """
    Make writer extension predictable and attach the right engine.
    - Lowercase xlsx/xlsm and use openpyxl.
    - For .xls or unknown junk, upgrade to .xlsx and use openpyxl.
    """
    root, ext = os.path.splitext(path)
    ext_low = ext.lower()
    if ext_low in ('.xlsx', '.xlsm'):
        return root + ext_low, {'engine': 'openpyxl'}
    if ext_low == '.xls':
        new_path = root + '.xlsx'
        dbg(f"[writer] upgrading output extension from {ext} to .xlsx")
        return new_path, {'engine': 'openpyxl'}
    # no or weird extension → force .xlsx
    new_path = (root if ext else path) + '.xlsx'
    dbg(f"[writer] forcing .xlsx for unknown ext {ext or '(none)'}")
    return new_path, {'engine': 'openpyxl'}

def normalise_date_any(val, date_format: str | None = None) -> pd.Timestamp:
    """Return pandas.Timestamp (UTC-naive) or NaT if invalid.
       If date_format is provided, parse strictly to that format; else auto-guess."""
    print(f"\n\n\n Date Format received: {date_format},, {val} \n\n\n")
    if val is None or (isinstance(val, float) and np.isnan(val)) or str(val).strip() == '':
        return pd.NaT

    s = str(val).strip()

    # Excel serial (integer days since 1899-12-30) – allowed in both modes
    if re.fullmatch(r'\d{1,6}', s):
        try:
            serial = int(s)
            if serial > 0:
                return EXCEL_EPOCH + pd.Timedelta(days=serial)
        except Exception:
            return pd.NaT

    # Strip timezone suffixes and unify separators
    s = re.sub(r'\s*\+\d{4}$', '', s).rstrip('Zz').strip()

    if date_format:
        # STRICT mode: parse exactly as given format
        s_norm = s.replace('/', '-').replace('.', '-')
        fmt_raw = date_format.strip().lower().replace('/', '-').replace('.', '-')

        # accept either python strptime tokens or simple tokens like yyyy-mm-dd, mm-dd-yy, etc.
        if '%' in fmt_raw:
            fmt_py = fmt_raw
        else:
            # replace longer tokens first to avoid overlaps
            fmt_py = (fmt_raw
                      .replace('yyyy', '%Y')
                      .replace('yy',   '%y')
                      .replace('mm',   '%m')
                      .replace('dd',   '%d'))
        try:
            dt = datetime.strptime(s_norm, fmt_py)
            print(f"\n\n\n Debug: Date parsed: {dt} \n\n\n")
            return pd.Timestamp(dt.date())
        except Exception:
            return pd.NaT

    # AUTO mode (original behavior)
    s_auto = s.replace('/', '-').replace('.', '-')
    for dayfirst in (True, False):
        try:
            dt = dparse.parse(s_auto, dayfirst=dayfirst, fuzzy=True, default=datetime(1900, 1, 1))
            return pd.Timestamp(dt.date())
        except Exception:
            continue

    return pd.NaT


def clean_billing_increment(val) -> str:
    """
    Normalize vendor increments with the rule:
      - If there are exactly two numbers -> keep as-is (no changes).
      - If there are 3+ numbers and any zeros -> drop all zeros, then take the first two remaining.
      - If only one number -> duplicate it (n/n).
      - If nothing useful remains (all zeros, or empty) -> return ''.

    Examples:
      "0/1/1"   -> "1/1"
      "1/0/60"  -> "1/60"
      "60/60"   -> "60/60"     (unchanged)
      "60"      -> "60/60"
      "0/0/0"   -> ""
    """
    if pd.isna(val):
        return ''
    nums = [int(n) for n in re.findall(r'\d+', str(val))]

    if not nums:
        return ''

    if len(nums) == 2:
        # leave exactly-two-value cases untouched (per your spec)
        return f"{nums[0]}/{nums[1]}"

    if len(nums) >= 3:
        # remove all zeros, then keep the first two remaining
        nz = [n for n in nums if n != 0]
        if len(nz) >= 2:
            return f"{nz[0]}/{nz[1]}"
        if len(nz) == 1:
            return f"{nz[0]}/{nz[0]}"
        return ''  # all zeros like "0/0/0"

    # len(nums) == 1
    return f"{nums[0]}/{nums[0]}"


def normalize_dates(df, column_name, date_format_email):
    """
    Normalize the date column to the required format and strip the time based on the date_format_email.
    """

    # Remove any time-related part after the date using regex
    print(f"\n\n\nBefore extracting date part:\n{df[column_name]}\n\n\n")

    # Replace . or / with -
    df[column_name] = df[column_name].str.replace(r'[./]', '-', regex=True)

    print(f"\n\n\nAfter replacing . or / with -:\n{df[column_name]}\n\n\n")

    df[column_name] = df[column_name].astype(str).str.extract(r'(\d{1,4}[-./]\d{1,2}[-./]\d{1,4})')[0]
   

    print(f"\n\n\nAfter extracting date part:\n{df[column_name]}\n\n\n")

    if date_format_email == 'MM-DD-YYYY':
        # Handle MM-DD-YYYY format, convert to YYYY-MM-DD
        df[column_name] = pd.to_datetime(df[column_name], format='%m-%d-%Y', errors='coerce')
        df[column_name] = df[column_name].dt.strftime('%Y-%m-%d')

    elif date_format_email == 'DD-MM-YYYY':
        # Handle DD-MM-YYYY format, convert to YYYY-MM-DD
        df[column_name] = pd.to_datetime(df[column_name], format='%d-%m-%Y', errors='coerce')
        df[column_name] = df[column_name].dt.strftime('%Y-%m-%d')

    elif date_format_email == 'YYYY-MM-DD':
        # Handle YYYY-MM-DD format, no conversion needed
        df[column_name] = pd.to_datetime(df[column_name], format='%Y-%m-%d', errors='coerce')
        df[column_name] = df[column_name].dt.strftime('%Y-%m-%d')

    else:
        # If the date format doesn't match any of the three options, raise an exception
        raise ValueError(f"Unsupported date format: {date_format_email}")

    return df



def load_clean_rates(path: str, output_path: str, sheet=None, date_format_email: str | None = None) -> pd.DataFrame:
    """
    Robust loader:
      1) Read raw grid (openpyxl for Excel; pandas for CSV/TXT)
      2) Detect header row (using your detect_header_row)
      3) Build a proper DataFrame with headers
      4) Canonicalize headers, trim notes/footer, select required cols
      5) Clean fields and write to Excel
    """
    if not os.path.exists(path):
        raise FileNotFoundError(f'File not found: {path}')

    # 1) raw grid
    raw = _read_raw_matrix(path, sheet=sheet)

    print(f"\n\n\nInitial raw data read from file:\n{raw}\n\n\n")
    # return raw

    print('\n\nDEBUG: Calling the detect header row function from load_clean_rates\n\n')
    # 2) detect header row in the raw grid
    header_row_idx = detect_header_row(raw)

    # 3) construct DF: header = that row; data = rows below it
    header_values = list(raw.iloc[header_row_idx].fillna('').astype(str))
    df = raw.iloc[header_row_idx+1:].copy()
    df.columns = header_values

    # drop columns/rows that are completely empty after slicing
    df.dropna(how="all", inplace=True)
    df.dropna(axis=1, how="all", inplace=True)
    df.reset_index(drop=True, inplace=True)

    # 4) canonicalize & trim
    df = _canonicalize_headers(df)
    df = _synthesize_billing_increment(df)
    df = trim_after_notes_and_strip_blank_above(df)

    is_js = "jerasoft" in str(path).lower()
    required_cols = list(REQUIRED_COLS)
    if is_js and "Dst Code Name" not in required_cols:
        required_cols = required_cols + ["Dst Code Name"]

    df = df[required_cols].copy()

    # 5) clean fields
    # Dst Code: keep as string, strip; drop truly empty codes
    df['Dst Code'] = df['Dst Code'].astype(str).str.strip()
    df = df[df['Dst Code'].ne("")]

    # Rate: strip symbols/spaces and coerce to float
    s = df['Rate'].astype(str).str.strip()
    s = (s
         .str.replace(r'[\$\£\€]', '', regex=True)
         .str.replace(r'\s+', '', regex=True))

 
    df['Rate'] = pd.to_numeric(s, errors='coerce')

    # Billing Increment: normalize to "x/y"
    df['Billing Increment'] = df['Billing Increment'].astype(str).str.strip().apply(clean_billing_increment)

    # Effective Date: robust parse (your helper returns Timestamp or NaT)
    # df['Effective Date'] = df['Effective Date'].apply(normalise_date_any)


    print(f"\n\n\nBefore normalizing Effective Date normalize_dates:\n{df['Effective Date']}\n\n\n")

    # If date_format_email is provided, normalize the dates
    if date_format_email:
        df = normalize_dates(df, 'Effective Date', date_format_email= date_format_email)


    # df['Effective Date'] = df['Effective Date'].apply(lambda v: normalise_date_any(v, date_format=date_format_email))

    print(f"\n\n\nAfter normalizing Effective Date:\n{df['Effective Date']}\n\n\n")

    df = expand_dst_code_rows(df)

    # Strip any non-digits left in Dst Code (kills '-', ';', spaces, exotic dashes, etc.)
    df['Dst Code'] = (
        df['Dst Code']
        .astype(str)
        .str.strip()
        .str.replace(r'\D+', '', regex=True)  # keep only 0-9
    )


    # finally, write the cleaned sheet
    out_path, writer_kwargs = _normalize_excel_writer_path(output_path)
    df.to_excel(out_path, index=False, **writer_kwargs)
    return df

# ──────────────────────────── quick test ─────────────────────────────────────
if __name__ == '__main__':
    PATH = r'C:\Users\User\OneDrive - Hayo Telecom, Inc\Documents\Work\Rate Sheet Automation\rate-sheet-automation\attachments\Book1.xlsx'
    OUT_PATH = r'C:\Users\User\OneDrive - Hayo Telecom, Inc\Documents\Work\Rate Sheet Automation\rate-sheet-automation\attachments\Book1_cleaned.xlsx'
    FILE_PATH = PATH
    OUTPUT_FILE_PATH = OUT_PATH 
    cleaned = load_clean_rates(FILE_PATH, OUTPUT_FILE_PATH, 0, date_format_email='DD-MM-YYYY')
   
    print('✅ Cleaned and saved.')