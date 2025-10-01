from openpyxl import load_workbook
from zipfile import ZipFile
from pathlib import Path
import os

PATH = r'attachments/rates_at_cimatelecom.com_20251001_003531/20250930224635_51098_24942.xlsx'

print("---- Basic file checks ----")
print("Exists:", os.path.exists(PATH), "Size (bytes):", os.path.getsize(PATH) if os.path.exists(PATH) else "n/a")
with open(PATH, 'rb') as f:
    sig = f.read(4)
print("ZIP signature (PK\\x03\\x04):", sig == b'PK\x03\x04', sig)

print("\n---- Open with openpyxl ----")
try:
    wb = load_workbook(PATH, read_only=True, data_only=True)
    print("wb.sheetnames:", wb.sheetnames)             # titles of normal worksheets
    print("#worksheets:", len(wb.worksheets))          # count of normal worksheets
    print("#_sheets total:", len(wb._sheets))          # includes chartsheets etc.

    # dump all sheet-like objects openpyxl found
    for sh in wb._sheets:
        print("  ", type(sh).__name__, getattr(sh, "title", "?"))

    # dump visibility/state of normal worksheets (if any)
    for ws in wb.worksheets:
        print("  worksheet:", ws.title, "state:", ws.sheet_state)
except Exception as e:
    print("openpyxl load error:", repr(e))

print("\n---- Inspect XLSX internals ----")
try:
    with ZipFile(PATH) as z:
        names = z.namelist()
        ws_files = [n for n in names if n.startswith("xl/worksheets/")]
        cs_files = [n for n in names if n.startswith("xl/chartsheets/")]
        print("worksheets xml:", ws_files)
        print("chartsheets  xml:", cs_files)
except Exception as e:
    print("zip inspection error:", repr(e))
